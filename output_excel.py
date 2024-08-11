import os
import subprocess
from pathlib import Path
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment


base_dir = Path(__file__).parent
excel_path = base_dir / 'chat_log.xlsx'


def is_output_open_excel() -> bool:
    """
    tell you if the Excel file is open or not
    :return: if it is open or not
    """
    #windows
    if os.name == "nt":
        try:
            with excel_path.open("r+b"):
                return False
        except PermissionError:
            return True
        except FileNotFoundError:
            return False
    #mac
    if os.name == "posix":
        pass


def load_or_create_workbook() -> tuple[openpyxl.Workbook,bool]:
    """
    read Excel file or create it if it does not exist
    :return: workbook object and flag if the file is created or not
    """
    # checking file existence
    if excel_path.exists():
        # read file and return it
        wb = openpyxl.load_workbook(excel_path)
        return wb, False

    else:
        # create file and return it
        wb = openpyxl.Workbook()
        return wb, True

def create_worksheet (title:str,target_workbook:openpyxl.Workbook,is_new:bool):
    """

    :param title: sheet title
    :param target_workbook: workbook object
    :param is_new: if workbook is created new or not
    :return: worksheet object
    """
    # removing the invalid title
    trimmed_title = trim_invalid_chars(title)
    if is_new:
        # getting active sheet
        target_worksheet = target_workbook.active
        target_worksheet.title = title

    else:
        # adding sheets
        target_worksheet = target_workbook.create_sheet(title=trimmed_title)
        target_workbook.move_sheet(target_worksheet,offset=-len(target_workbook.worksheets)+1)
        target_workbook.active = target_worksheet
    return target_worksheet

def trim_invalid_chars(title:str) -> str:
    """
    remove the invalid character from the sheet
    :param title: sheet title
    :return: sheet title after removing
    """
    new_title = title
    invalid_chars = ["/","\\","?","*","[","]"]
    for char in invalid_chars:
        new_title = new_title.replace(char,"")
    return new_title

def header_formatting(target_worksheet):
    """
    change the header of worksheet
    :param ws: worksheet to export
    :return:
    """
    # write date on A1
    datetime_cell = target_worksheet["A1"]
    target_worksheet.value = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    target_worksheet["A1"].font = Font(name="Times New Roman")

    # set character on A2 and B2
    role_header_cell = target_worksheet["A2"]
    content_header_cell = target_worksheet["B2"]

    # set values on cell
    role_header_cell.value = "Role"
    content_header_cell.value = "Description"

    # set font
    white_color = "FFFFFF"
    header_font_style = Font(name="Times New Roman", bold=True, color=white_color)
    role_header_cell.font = header_font_style
    content_header_cell.font = header_font_style

    # set cell color
    excel_green = "156B31"
    header_color = PatternFill(fill_type="solid", fgColor=excel_green)
    role_header_cell.fill = header_color
    content_header_cell.fill = header_color

    # set cell width
    # ws.columns.dimensions["A"].width = 22
    # ws.columns.dimensions["B"].width = 168
    target_worksheet.column_dimensions["A"].width = 22
    target_worksheet.column_dimensions["B"].width = 165

def write_chat_log(target_worksheet,chat_log: list[dict]):
    """
    set the format for the chat history
    :param ws: worksheet to export
    :param chat_log: chat history
    """
    row_height_adjustment_standard = 18
    font_style = Font(name="Times New Roman", size=10)
    light_gray = "d9d9d9"
    assistant_color = PatternFill(fill_type="solid",fgColor="d9d9d9")

    write_start_low = 3
    for row_number, content in enumerate(chat_log, 3):
        cell_role, cell_content = target_worksheet[f"A{row_number}"], target_worksheet[f"B{row_number}"]

        # write down on the cell
        cell_role.value = content["role"]
        cell_content.value = content["content"]

        # cell put space
        cell_content.alignment = Alignment(wrapText=True)

        # adjust culum height
        adjusted_row_height = len(content["content"].split("\n"))*row_height_adjustment_standard
        target_worksheet.row_dimensions[row_number].height = adjusted_row_height

        # setting format
        cell_role.font = font_style
        cell_content.font = font_style
        if content["role"] == "assistant":
            cell_role.fill = assistant_color
            cell_content.fill = assistant_color

def open_workbook():
    """open excel file"""
    # windows
    if os.name == "nt":
        os.system(f"start {excel_path}")

    # mac
    if os.name == "posix":
        os.system(f"open{excel_path}")

def output_excel(chat_log:list[dict], chat_summary:str):
    """
    entry point to write the chat history on chat_hitstory.xlsx
    :param chat_log: chat history
    :param chat_summary: chat summary
    :return:
    """

    workbook, is_created = load_or_create_workbook()
    worksheet = create_worksheet(chat_summary,workbook,is_created)
    header_formatting(target_worksheet=worksheet)
    write_chat_log(target_worksheet=worksheet, chat_log=chat_log)
    workbook.save(excel_path)
    workbook.close()


log = [
        {"role": "user", "content": "Hello"},
        {"role": "assistant", "content": "AI assistant"},
        {"role": "user", "content": "how are you?"},
        {"role": "assistant", "content": "I am fine\n tha\n nks,\nyou?\n not\n much?\n huh?\n really?\n how you doing\n these days"}
    ]

output_excel(log,"test/\\?*")